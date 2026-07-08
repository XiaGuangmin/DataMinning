#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""科技主线股票池盘后扫描

数据源：腾讯财经（主）、新浪财经（备），不依赖东方财富接口。
输出：output/scan_<日期>.xlsx（4个Sheet）+ output/summary_<日期>.md（摘要）。

用法：python3 scan.py
非交易日（当日无K线数据）时打印 NON_TRADING_DAY 并以退出码 0 结束。
"""

import json
import os
import re
import ssl
import sys
import time
import urllib.request
from datetime import datetime, timedelta, timezone

# 北京时区：优先用系统时区库，缺失（如 Windows 未装 tzdata）时退回固定 UTC+8
try:
    from zoneinfo import ZoneInfo
    CN_TZ = ZoneInfo("Asia/Shanghai")
except Exception:  # noqa: BLE001
    CN_TZ = timezone(timedelta(hours=8))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
POOL_FILE = os.path.join(BASE_DIR, "pool.json")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# 池内常用简称 -> 行情接口注册名，避免误报代码校验警告
KNOWN_ALIASES = {
    "韦尔股份": "豪威集团",
    "华虹公司": "华虹宏力",
}

# ---------------------------------------------------------------- HTTP

def _build_opener():
    # 本机直连即可；若处于需要代理/自定义 CA 的环境，走标准环境变量
    cafile = os.environ.get("SSL_CERT_FILE") or os.environ.get("CURL_CA_BUNDLE")
    ctx = ssl.create_default_context(cafile=cafile) if cafile else ssl.create_default_context()
    return urllib.request.build_opener(
        urllib.request.ProxyHandler(urllib.request.getproxies()),
        urllib.request.HTTPSHandler(context=ctx),
    )

_OPENER = _build_opener()


def http_get(url, headers=None, timeout=15, retries=3):
    last_err = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers or {})
            with _OPENER.open(req, timeout=timeout) as resp:
                return resp.read()
        except Exception as e:  # noqa: BLE001
            last_err = e
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"请求失败 {url}: {last_err}")


def market_prefix(code):
    return "sh" if code.startswith(("6", "9")) else "sz"


# ---------------------------------------------------------------- 行情

def fetch_quotes_tencent(codes):
    """批量实时行情（盘后为收盘快照）。返回 {code: dict}。"""
    result = {}
    for i in range(0, len(codes), 50):
        batch = codes[i:i + 50]
        q = ",".join(market_prefix(c) + c for c in batch)
        raw = http_get(f"https://qt.gtimg.cn/q={q}").decode("gbk", errors="replace")
        for line in raw.strip().split(";"):
            line = line.strip()
            m = re.match(r'v_(?:sh|sz)(\d{6})="(.*)"', line)
            if not m:
                continue
            code, body = m.group(1), m.group(2)
            f = body.split("~")
            if len(f) < 50:
                continue
            def num(idx):
                try:
                    return float(f[idx])
                except (ValueError, IndexError):
                    return None
            result[code] = {
                "name": f[1].strip(),
                "close": num(3),
                "prev_close": num(4),
                "pct": num(32),
                "volume": num(36),        # 手
                "turnover": num(38),      # 换手率 %
                "vol_ratio": num(49),     # 量比
                "source": "tencent",
            }
    return result


def fetch_quotes_sina(codes):
    """新浪备用行情。无换手率/量比（量比后续用K线计算）。"""
    result = {}
    headers = {"Referer": "https://finance.sina.com.cn"}
    for i in range(0, len(codes), 50):
        batch = codes[i:i + 50]
        q = ",".join(market_prefix(c) + c for c in batch)
        raw = http_get(f"https://hq.sinajs.cn/list={q}", headers=headers).decode("gbk", errors="replace")
        for line in raw.strip().split("\n"):
            m = re.match(r'var hq_str_(?:sh|sz)(\d{6})="(.*)";', line.strip())
            if not m or not m.group(2):
                continue
            code, f = m.group(1), m.group(2).split(",")
            if len(f) < 32:
                continue
            close, prev = float(f[3]), float(f[2])
            result[code] = {
                "name": f[0].strip(),
                "close": close,
                "prev_close": prev,
                "pct": round((close / prev - 1) * 100, 2) if prev else None,
                "volume": float(f[8]) / 100,  # 股 -> 手
                "turnover": None,
                "vol_ratio": None,
                "source": "sina",
            }
    return result


def fetch_kline(code, days=160):
    """日K线 [(date, open, close, high, low, volume手), ...]，前复权。"""
    sym = market_prefix(code) + code
    url = (f"https://web.ifzq.gtimg.cn/appstock/app/fqkline/get"
           f"?param={sym},day,,,{days},qfq")
    data = json.loads(http_get(url))
    node = data["data"][sym]
    rows = node.get("qfqday") or node.get("day") or []
    out = []
    for r in rows:
        out.append((r[0], float(r[1]), float(r[2]), float(r[3]), float(r[4]), float(r[5])))
    return out


# ---------------------------------------------------------------- 指标

def ema(values, n):
    out, k = [], 2 / (n + 1)
    for i, v in enumerate(values):
        out.append(v if i == 0 else v * k + out[-1] * (1 - k))
    return out


def macd(closes):
    dif = [a - b for a, b in zip(ema(closes, 12), ema(closes, 26))]
    dea = ema(dif, 9)
    bar = [2 * (a - b) for a, b in zip(dif, dea)]
    return dif, dea, bar


def kdj(highs, lows, closes, n=9):
    ks, ds, js = [], [], []
    k = d = 50.0
    for i in range(len(closes)):
        lo, hi = min(lows[max(0, i - n + 1):i + 1]), max(highs[max(0, i - n + 1):i + 1])
        rsv = (closes[i] - lo) / (hi - lo) * 100 if hi > lo else 50.0
        k = 2 / 3 * k + 1 / 3 * rsv
        d = 2 / 3 * d + 1 / 3 * k
        ks.append(k); ds.append(d); js.append(3 * k - 2 * d)
    return ks, ds, js


def rsi(closes, n=14):
    out, avg_g, avg_l = [50.0], 0.0, 0.0
    for i in range(1, len(closes)):
        chg = closes[i] - closes[i - 1]
        g, l = max(chg, 0), max(-chg, 0)
        if i <= n:
            avg_g += g / n
            avg_l += l / n
        else:
            avg_g = (avg_g * (n - 1) + g) / n
            avg_l = (avg_l * (n - 1) + l) / n
        out.append(100 - 100 / (1 + avg_g / avg_l) if avg_l > 0 else 100.0)
    return out


def cross_days_ago(fast, slow, direction, lookback=3):
    """fast 上穿(up)/下穿(down) slow 发生在几天前（0=今天）；无则 None。"""
    for ago in range(lookback):
        i = len(fast) - 1 - ago
        if i < 1:
            break
        if direction == "up" and fast[i] > slow[i] and fast[i - 1] <= slow[i - 1]:
            return ago
        if direction == "down" and fast[i] < slow[i] and fast[i - 1] >= slow[i - 1]:
            return ago
    return None


def find_pivots(values, mode, w=3):
    """局部极值下标（mode=min/max），窗口 ±w。"""
    idx = []
    for i in range(w, len(values) - w):
        seg = values[i - w:i + w + 1]
        if mode == "min" and values[i] == min(seg):
            idx.append(i)
        elif mode == "max" and values[i] == max(seg):
            idx.append(i)
    return idx


def divergence(closes, dif, kind, window=60, recent=10):
    """MACD 背离检测。kind=bottom/top。返回 (bool, 强度) 。

    bottom: 近 window 内最后两个价格波谷，价新低而 DIF 未新低。
    强度按 DIF 抬升/回落幅度相对值分档：>30% 强，>10% 中，其余弱。
    """
    start = max(0, len(closes) - window)
    seg_c = closes[start:]
    seg_d = dif[start:]
    pivots = find_pivots(seg_c, "min" if kind == "bottom" else "max")
    if len(pivots) < 2:
        return False, ""
    p1, p2 = pivots[-2], pivots[-1]
    if len(seg_c) - 1 - p2 > recent:  # 最近一个极值太久远，不算当前背离
        return False, ""
    if kind == "bottom":
        ok = seg_c[p2] < seg_c[p1] and seg_d[p2] > seg_d[p1]
    else:
        ok = seg_c[p2] > seg_c[p1] and seg_d[p2] < seg_d[p1]
    if not ok:
        return False, ""
    rel = abs(seg_d[p2] - seg_d[p1]) / (abs(seg_d[p1]) + 1e-9)
    return True, "强" if rel > 0.3 else ("中" if rel > 0.1 else "弱")


# ---------------------------------------------------------------- 个股分析

def analyze_stock(quote, kline):
    """返回筛选与技术指标结果 dict。kline 需含当日。"""
    closes = [r[2] for r in kline]
    highs = [r[3] for r in kline]
    lows = [r[4] for r in kline]
    vols = [r[5] for r in kline]
    n = len(closes)
    r = {"data_ok": n >= 70}
    if not r["data_ok"]:
        return r

    pct = (closes[-1] / closes[-2] - 1) * 100 if n >= 2 else None
    gain5 = (closes[-1] / closes[-6] - 1) * 100 if n >= 6 else None
    ma20 = sum(closes[-20:]) / 20
    ma60 = sum(closes[-60:]) / 60
    vol_ratio = quote.get("vol_ratio")
    if not vol_ratio and n >= 6:
        avg5 = sum(vols[-6:-1]) / 5
        vol_ratio = round(vols[-1] / avg5, 2) if avg5 else None
    turnover = quote.get("turnover")

    # 刚站上20/60日线：今日在均线上方，且近3日内有收盘价在其下方
    def just_crossed(ma_n):
        ma_series = [sum(closes[i - ma_n + 1:i + 1]) / ma_n for i in range(ma_n - 1, n)]
        cs = closes[ma_n - 1:]
        if cs[-1] <= ma_series[-1]:
            return False
        return any(cs[-1 - a] <= ma_series[-1 - a] for a in range(1, 4) if len(cs) > a)

    cross20, cross60 = just_crossed(20), just_crossed(60)
    # 横盘后突破：前3日之前的20日收盘振幅≤10%，今收突破该区间高点
    base = closes[-23:-3]
    breakout = False
    if len(base) == 20 and min(base) > 0:
        breakout = (max(base) / min(base) - 1) <= 0.10 and closes[-1] > max(base)

    reasons = []
    quant_ok = (turnover is not None and 3 <= turnover <= 8
                and vol_ratio is not None and 1.5 <= vol_ratio <= 3
                and gain5 is not None and 5 <= gain5 <= 15)
    if quant_ok:
        reasons.append(f"换手{turnover:.1f}%/量比{vol_ratio:.2f}/5日涨{gain5:.1f}%")
        if breakout:
            reasons.append("横盘突破")
        if cross20:
            reasons.append("刚站上20日线")
        if cross60:
            reasons.append("刚站上60日线")
    launch = quant_ok and (breakout or cross20 or cross60)

    accel = ""
    accel_why = []
    if turnover is not None and turnover > 15:
        accel, accel_why = "红色", [f"换手{turnover:.1f}%>15%"]
    else:
        if turnover is not None and 10 <= turnover <= 15:
            accel_why.append(f"换手{turnover:.1f}%")
        if vol_ratio is not None and 3 < vol_ratio <= 5:
            accel_why.append(f"量比{vol_ratio:.2f}")
        if gain5 is not None and 15 < gain5 <= 25:
            accel_why.append(f"5日涨{gain5:.1f}%")
        if accel_why:
            accel = "黄色"

    # ---- 技术指标
    dif, dea, _bar = macd(closes)
    ks, ds, js = kdj(highs, lows, closes)
    rs = rsi(closes)

    gc_ago = cross_days_ago(dif, dea, "up")
    dc_ago = cross_days_ago(dif, dea, "down")
    macd_gc = f"金叉(零轴{'上' if dif[-1 - gc_ago] >= 0 else '下'}方)" if gc_ago is not None else ""
    macd_dc = f"死叉(零轴{'上' if dif[-1 - dc_ago] >= 0 else '下'}方)" if dc_ago is not None else ""
    bot_div, bot_str = divergence(closes, dif, "bottom")
    top_div, top_str = divergence(closes, dif, "top")

    kdj_gc = cross_days_ago(ks, ds, "up") is not None and min(js[-5:]) < 20
    kdj_dc = cross_days_ago(ks, ds, "down") is not None and max(js[-5:]) > 80
    rsi_rebound = len(rs) > 4 and min(rs[-4:-1]) < 30 and rs[-1] > rs[-2]
    rsi_fall = len(rs) > 4 and max(rs[-4:-1]) > 70 and rs[-1] < rs[-2]

    buys = []
    if bot_div:
        buys.append(f"MACD底背离({bot_str})")
    if macd_gc:
        buys.append(f"MACD{macd_gc}")
    if kdj_gc:
        buys.append("KDJ超卖金叉")
    if rsi_rebound:
        buys.append("RSI超卖反弹")
    sells = []
    if top_div:
        sells.append(f"MACD顶背离({top_str})")
    if macd_dc:
        sells.append(f"MACD{macd_dc}")
    if kdj_dc:
        sells.append("KDJ超买死叉")
    if rsi_fall:
        sells.append("RSI超买回落")

    resonance = ""
    if len(buys) >= 2:
        resonance = "强买点共振"
    if len(sells) >= 2:
        resonance = (resonance + "+" if resonance else "") + "强卖点共振"

    if buys and not sells:
        judge = "偏买点"
    elif sells and not buys:
        judge = "偏卖点"
    elif buys and sells:
        judge = "多空信号并存，观望"
    else:
        judge = "无明确信号"

    macd_status = macd_gc or macd_dc or ("DIF>DEA多头" if dif[-1] > dea[-1] else "DIF<DEA空头")

    r.update({
        "close": closes[-1],
        "pct": pct, "gain5": gain5, "ma20": ma20, "ma60": ma60,
        "above_ma20": closes[-1] > ma20, "above_ma60": closes[-1] > ma60,
        "vol_ratio": vol_ratio, "turnover": turnover,
        "breakout": breakout, "cross20": cross20, "cross60": cross60,
        "launch": launch, "launch_reasons": reasons,
        "accel": accel, "accel_why": accel_why,
        "dif": dif[-1], "dea": dea[-1], "macd_status": macd_status,
        "k": ks[-1], "d": ds[-1], "j": js[-1], "rsi": rs[-1],
        "buys": buys, "sells": sells, "resonance": resonance, "judge": judge,
        "kline_dates": [row[0] for row in kline],
    })
    return r


# ---------------------------------------------------------------- 主流程

def main():
    # SCAN_DATE 可覆盖交易日基准（格式 YYYY-MM-DD），用于回补历史或测试
    today = os.environ.get("SCAN_DATE") or datetime.now(CN_TZ).strftime("%Y-%m-%d")

    with open(POOL_FILE, encoding="utf-8") as f:
        pool = json.load(f)

    stocks = []  # (category, name, code, is_candidate, cand_meta)
    for cat, items in pool["fixed_pool"].items():
        for s in items:
            stocks.append((cat, s["name"], s["code"], False, None))
    for s in pool["candidate_pool"]:
        stocks.append(("候选池", s["name"], s["code"], True, s))
    codes = [s[2] for s in stocks]

    # 交易日判断：基准股当日是否有K线
    bench = fetch_kline("600519", days=5)
    if not bench or bench[-1][0] != today:
        print(f"NON_TRADING_DAY {today} 最新K线日期={bench[-1][0] if bench else '无'}")
        return 0

    warnings = []
    try:
        quotes = fetch_quotes_tencent(codes)
        if len(quotes) < len(codes) * 0.8:
            raise RuntimeError(f"腾讯行情仅返回 {len(quotes)}/{len(codes)} 只")
    except Exception as e:  # noqa: BLE001
        warnings.append(f"腾讯行情失败({e})，切换新浪备用源")
        quotes = fetch_quotes_sina(codes)

    results = {}
    for cat, name, code, _isc, _cm in stocks:
        q = quotes.get(code)
        if not q:
            warnings.append(f"{name}({code}) 无行情数据")
            continue
        api_name = re.sub(r"^(XD|XR|DR|ST|\*ST|N)\s*", "", q["name"]).replace(" ", "")
        if api_name and api_name != name and api_name != KNOWN_ALIASES.get(name):
            warnings.append(f"代码校验:{code} 池内名[{name}] 接口名[{q['name']}]")
        try:
            kl = fetch_kline(code)
            results[code] = analyze_stock(q, kl)
            results[code]["quote"] = q
        except Exception as e:  # noqa: BLE001
            warnings.append(f"{name}({code}) K线获取失败: {e}")
        time.sleep(0.12)

    # 板块统计（仅固定池）
    sectors = {}
    for cat, name, code, isc, _cm in stocks:
        if isc or code not in results or not results[code].get("data_ok"):
            continue
        pct = results[code].get("pct") or 0
        sectors.setdefault(cat, []).append((name, pct))
    thr = pool["meta"].get("sector_alert_threshold_pct", 2.0)
    sector_rows = []
    for cat, lst in sectors.items():
        avg = sum(p for _, p in lst) / len(lst)
        ups = sum(1 for _, p in lst if p > 0)
        top = sorted(lst, key=lambda x: -x[1])[:3]
        sector_rows.append({
            "cat": cat, "avg": avg, "ups": ups, "total": len(lst),
            "alert": avg >= thr,
            "top": "、".join(f"{n}({p:+.1f}%)" for n, p in top),
        })
    sector_rows.sort(key=lambda x: -x["avg"])

    # 候选池到期评估（观察满3个交易日）
    cand_evals = []
    for cat, name, code, isc, cm in stocks:
        if not isc or not cm or cm.get("status") != "observing" or not cm.get("added_date"):
            continue
        rr = results.get(code)
        if not rr or not rr.get("data_ok"):
            continue
        dates = rr["kline_dates"]
        obs_days = sum(1 for d in dates if d > cm["added_date"])
        if obs_days >= 3:
            idx = next((i for i, d in enumerate(dates) if d > cm["added_date"]), None)
            gain = None
            if idx is not None and idx > 0:
                closes_all = None  # gain 由 quote 近似：用5日涨幅替代观察期涨幅不准，直接算
            # 观察期涨幅 = 今收 / 入池前收 - 1
            kl = fetch_kline(code, days=obs_days + 10)
            base_close = next((r_[2] for r_ in reversed(kl) if r_[0] <= cm["added_date"]), None)
            if base_close:
                gain = (kl[-1][2] / base_close - 1) * 100
            good = (gain is not None and gain > 0
                    and rr.get("turnover") is not None and 3 <= rr["turnover"] <= 15)
            cand_evals.append({
                "name": name, "code": code, "obs_days": obs_days, "gain": gain,
                "advice": "建议晋升固定池" if good else "建议移出候选池",
            })

    launch_list = [(c, n, code) for c, n, code, _i, _m in stocks
                   if code in results and results[code].get("launch")]
    accel_list = [(c, n, code, results[code]["accel"]) for c, n, code, _i, _m in stocks
                  if code in results and results[code].get("accel")]
    signal_codes = {code for _c, _n, code in launch_list} | {x[2] for x in accel_list}

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    xlsx = os.path.join(OUTPUT_DIR, f"scan_{today}.xlsx")
    write_excel(xlsx, today, stocks, results, launch_list, accel_list,
                signal_codes, sector_rows, thr)

    summary = build_summary(today, stocks, results, launch_list, accel_list,
                            signal_codes, sector_rows, thr, cand_evals, warnings)
    md = os.path.join(OUTPUT_DIR, f"summary_{today}.md")
    with open(md, "w", encoding="utf-8") as f:
        f.write(summary)

    pool["meta"]["last_run"] = today
    with open(POOL_FILE, "w", encoding="utf-8") as f:
        json.dump(pool, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"OK 扫描完成 {today}  Excel: {xlsx}  摘要: {md}")
    print("=" * 40)
    print(summary)
    return 0


# ---------------------------------------------------------------- 输出

def write_excel(path, today, stocks, results, launch_list, accel_list,
                signal_codes, sector_rows, thr):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    green = PatternFill("solid", fgColor="C6EFCE")

    def header(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = bold

    # Sheet1 全量数据
    ws = wb.active
    ws.title = "全量数据"
    header(ws, ["分类", "名称", "代码", "收盘价", "涨跌幅%", "换手率%", "量比",
                "近5日涨幅%", "站上MA20", "站上MA60", "横盘突破", "信号标记"])
    for cat, name, code, _i, _m in stocks:
        r = results.get(code)
        if not r or not r.get("data_ok"):
            ws.append([cat, name, code] + ["数据缺失"] * 9)
            continue
        q = r["quote"]
        mark = ("启动信号" if r["launch"] else "") + \
               (("/" if r["launch"] and r["accel"] else "") +
                (f"{r['accel']}加速" if r["accel"] else ""))
        row = [cat, name, code, r["close"], r["pct"], r["turnover"],
               round(r["vol_ratio"], 2) if r["vol_ratio"] else None,
               round(r["gain5"], 2) if r["gain5"] is not None else None,
               "是" if r["above_ma20"] else "否",
               "是" if r["above_ma60"] else "否",
               "是" if r["breakout"] else "否", mark or "无"]
        ws.append(row)
        cell = ws.cell(row=ws.max_row, column=12)
        if r["accel"] == "红色":
            cell.fill = red
        elif r["accel"] == "黄色":
            cell.fill = yellow
        elif r["launch"]:
            cell.fill = green

    # Sheet2 信号汇总与操作建议
    ws = wb.create_sheet("信号汇总与操作建议")
    header(ws, ["类型", "分类", "名称", "代码", "理由", "技术指标状态", "操作建议"])
    for cat, name, code in launch_list:
        r = results[code]
        ws.append(["启动信号", cat, name, code, "；".join(r["launch_reasons"]),
                   f"{r['macd_status']}；{r['judge']}",
                   "初步启动，可小仓位试探，跌破20日线止损"])
        ws.cell(row=ws.max_row, column=1).fill = green
    for cat, name, code, level in accel_list:
        r = results[code]
        advice = ("过热风险大，不追高，持仓考虑分批兑现" if level == "红色"
                  else "趋势加速中，持仓可持有但设移动止盈，空仓不重仓追")
        ws.append([f"{level}加速", cat, name, code, "；".join(r["accel_why"]),
                   f"{r['macd_status']}；{r['judge']}", advice])
        ws.cell(row=ws.max_row, column=1).fill = red if level == "红色" else yellow

    # Sheet3 技术指标监测（仅信号个股）
    ws = wb.create_sheet("技术指标监测")
    header(ws, ["名称", "代码", "MACD状态", "DIF", "DEA", "K", "D", "J", "RSI",
                "买点信号", "卖点信号", "共振类型", "买卖点判断"])
    for cat, name, code, _i, _m in stocks:
        if code not in signal_codes:
            continue
        r = results[code]
        ws.append([name, code, r["macd_status"],
                   round(r["dif"], 3), round(r["dea"], 3),
                   round(r["k"], 1), round(r["d"], 1), round(r["j"], 1),
                   round(r["rsi"], 1),
                   "、".join(r["buys"]) or "无", "、".join(r["sells"]) or "无",
                   r["resonance"] or "无", r["judge"]])

    # Sheet4 板块异动分析
    ws = wb.create_sheet("板块异动分析")
    header(ws, ["细分方向", "平均涨跌幅%", "上涨家数", "总数", f"异动(≥{thr}%)", "涨幅前三"])
    for s in sector_rows:
        ws.append([s["cat"], round(s["avg"], 2), s["ups"], s["total"],
                   "是" if s["alert"] else "否", s["top"]])
        if s["alert"]:
            ws.cell(row=ws.max_row, column=5).fill = yellow

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max(len(str(c.value or "")) for c in col) + 2
            sheet.column_dimensions[col[0].column_letter].width = min(width * 1.6, 50)
    wb.save(path)


def build_summary(today, stocks, results, launch_list, accel_list,
                  signal_codes, sector_rows, thr, cand_evals, warnings):
    lines = [f"# 科技盘后扫描摘要 {today}", ""]

    lines.append("## 一、启动信号个股")
    if launch_list:
        for cat, name, code in launch_list:
            r = results[code]
            lines.append(f"- **{name}**({code}，{cat})：{'；'.join(r['launch_reasons'])}；"
                         f"{r['macd_status']}，{r['judge']}")
    else:
        lines.append("- 今日无满足全部启动条件的个股")
    lines.append("")

    lines.append("## 二、加速警告")
    reds = [x for x in accel_list if x[3] == "红色"]
    yels = [x for x in accel_list if x[3] == "黄色"]
    if reds:
        lines.append("**红色加速（过热，警惕兑现）：**")
        for cat, name, code, _l in reds:
            lines.append(f"- {name}({code}，{cat})：{'；'.join(results[code]['accel_why'])}")
    if yels:
        lines.append("**黄色加速（趋势加速，注意节奏）：**")
        for cat, name, code, _l in yels:
            lines.append(f"- {name}({code}，{cat})：{'；'.join(results[code]['accel_why'])}")
    if not accel_list:
        lines.append("- 今日无加速警告个股")
    lines.append("")

    lines.append(f"## 三、板块异动（平均涨幅≥{thr}%触发）")
    alerts = [s for s in sector_rows if s["alert"]]
    for s in sector_rows[:5]:
        flag = " ⚠️异动" if s["alert"] else ""
        lines.append(f"- {s['cat']}：平均{s['avg']:+.2f}%，{s['ups']}/{s['total']}上涨，"
                     f"领涨 {s['top']}{flag}")
    if alerts:
        lines.append("（板块异动触发：可按启动标准在该板块内筛选非池内候选股补入候选池）")
    lines.append("")

    lines.append("## 四、技术指标共振")
    reso = [(n, code) for _c, n, code, _i, _m in stocks
            if code in signal_codes and results.get(code, {}).get("resonance")]
    if reso:
        for n, code in reso:
            r = results[code]
            lines.append(f"- **{n}**({code})：{r['resonance']}——买点[{('、'.join(r['buys'])) or '无'}] "
                         f"卖点[{('、'.join(r['sells'])) or '无'}]")
    else:
        lines.append("- 信号个股中今日无2个及以上指标共振")
    lines.append("")

    if cand_evals:
        lines.append("## 五、候选池到期评估（观察满3日）")
        for c in cand_evals:
            g = f"{c['gain']:+.1f}%" if c["gain"] is not None else "N/A"
            lines.append(f"- {c['name']}({c['code']})：观察{c['obs_days']}日，期间涨幅{g} → **{c['advice']}**")
        lines.append("")

    if warnings:
        lines.append("## 数据警告")
        for w in warnings[:15]:
            lines.append(f"- {w}")
        lines.append("")

    lines.append("> 以上为程序化扫描结果，仅供盘后复盘参考，不构成投资建议。")
    return "\n".join(lines)


if __name__ == "__main__":
    sys.exit(main())
